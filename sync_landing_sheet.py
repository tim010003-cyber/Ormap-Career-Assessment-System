"""Sync 腳本：把 xlsx 的 04_Landing_Page文案 sheet 轉成 twa-landing.json

用法:
    python3 sync_landing_sheet.py [--xlsx PATH] [--out PATH]

預設:
    --xlsx data/TWA_完整資料庫_主檔.xlsx
    --out  data/twa-landing.json
"""
import argparse, json, sys
from pathlib import Path
from openpyxl import load_workbook

SHEET_NAME = '04_Landing_Page文案'

# content 欄的值轉換規則
def _cast(val):
    if val is None: return ''
    if isinstance(val, str):
        low = val.strip().lower()
        if low == 'true':  return True
        if low == 'false': return False
    return val

def xlsx_to_landing_json(xlsx_path: Path) -> dict:
    wb = load_workbook(xlsx_path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f'找不到 sheet: {SHEET_NAME}')
    ws = wb[SHEET_NAME]

    obj = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        # 只取前 5 欄
        section, item, key, content, _note = (row + (None,)*5)[:5]
        if not section or not key:  # 跳過空列
            continue
        content = _cast(content)
        sec = obj.setdefault(section, {})

        # 單一值（item 為 "-" 或空）
        if not item or item == '-':
            sec[key] = content
            continue

        # group_X_step_Y  (SOP 的 step)
        if item.startswith('group_') and '_step_' in item:
            gkey, skey = item.split('_step_', 1)
            gs = sec.setdefault('_groups', {})
            g = gs.setdefault(gkey, {})
            g.setdefault('_steps', {}).setdefault(skey, {})[key] = content
            continue

        # group_X  (SOP 的 group meta)
        if item.startswith('group_'):
            sec.setdefault('_groups', {}).setdefault(item, {})[key] = content
            continue

        # card_N  (motivation / reflection)
        if item.startswith('card_'):
            sec.setdefault('_cards', {}).setdefault(item, {})[key] = content
            continue

        # paragraph_N  (valueIntro)
        if item.startswith('paragraph_'):
            sec.setdefault('_paragraphs', {})[item] = content
            continue

        print(f'[warn] row {row_idx}: 無法識別 item={item!r}，已略過', file=sys.stderr)

    # 把 _groups / _cards / _paragraphs 整理為 list
    for sec_name, sec in obj.items():
        if '_cards' in sec:
            sec['cards'] = [sec['_cards'][k] for k in sorted(sec['_cards'].keys())]
            del sec['_cards']
        if '_groups' in sec:
            grp_dict = sec['_groups']
            for gk in grp_dict:
                if '_steps' in grp_dict[gk]:
                    step_dict = grp_dict[gk]['_steps']
                    grp_dict[gk]['steps'] = [step_dict[k] for k in sorted(step_dict.keys())]
                    del grp_dict[gk]['_steps']
            sec['groups'] = [grp_dict[k] for k in sorted(grp_dict.keys())]
            del sec['_groups']
        if '_paragraphs' in sec:
            p_dict = sec['_paragraphs']
            # 只保留非空 paragraph
            sec['paragraphs'] = [p_dict[k] for k in sorted(p_dict.keys()) if p_dict.get(k)]
            del sec['_paragraphs']

    return obj

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--xlsx', default='data/TWA_完整資料庫_主檔.xlsx')
    ap.add_argument('--out',  default='data/twa-landing.json')
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx)
    out_path = Path(args.out)
    if not xlsx_path.exists():
        print(f'錯誤：找不到 {xlsx_path}', file=sys.stderr); sys.exit(1)

    data = xlsx_to_landing_json(xlsx_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')
    print(f'✓ 寫入 {out_path}')
    print(f'  sections: {list(data.keys())}')

if __name__ == '__main__':
    main()
