#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
sync_indicators_sheet.py
-------------------------
把 TWA_完整資料庫_主檔.xlsx 的「02_環境指標清單」sheet
同步回 data/twa-matrix.json。

設計原則：
- indicators 陣列（text / label / weight / hover / idx）完全從 xlsx 覆蓋。
- 頂層欄位 `v`（價值分類）、`d`（需求描述）、`n`（需求名稱）保留現有 JSON
  不動（xlsx 沒有這些資料，不能從 xlsx 覆蓋）。
- 如果 xlsx 的 need 名稱與 JSON 的 n 對不上，印出警告但不阻斷。

用法：
    python sync_indicators_sheet.py
    python sync_indicators_sheet.py --xlsx 路徑.xlsx --json 路徑.json
    python sync_indicators_sheet.py --dry-run    # 不寫檔，只看會變什麼

預設路徑：
    --xlsx  data/TWA_完整資料庫_主檔.xlsx
    --json  data/twa-matrix.json
"""

import argparse
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("[錯誤] 缺少 openpyxl，請先安裝：pip install openpyxl")


# ---------------------------------------------------------------------------
# 設定
# ---------------------------------------------------------------------------

SHEET_NAME = "02_環境指標清單"
HEADER_ROW_COUNT = 2   # 第 1 列是說明、第 2 列是欄位標題、資料從第 3 列開始
COL_NEED_ID = 1        # A：need_id（0-19）
COL_NEED = 2           # B：心理需求名稱（參考用）
COL_INDICATOR_IDX = 3  # C：indicator_index
COL_WEIGHT = 4         # D：權重分數
COL_HOVER = 5          # E：調整理由與場景定義（Hover 說明文案）
COL_INDICATOR = 6      # F：indicator（指標名稱）
COL_LABEL = 7          # G：UI 顯示標籤（卡片標題）

DEFAULT_XLSX = Path("data/TWA_完整資料庫_主檔.xlsx")
DEFAULT_JSON = Path("data/twa-matrix.json")


# ---------------------------------------------------------------------------
# 工具函式
# ---------------------------------------------------------------------------

def clean_text(value):
    """清理 xlsx 讀進來的字串：去頭尾空白、None → ''。"""
    if value is None:
        return ""
    return str(value).strip()


def read_xlsx_indicators(xlsx_path):
    """讀 xlsx 並回傳 {need_id: [indicator_dict, ...]} 與 {need_id: need_name}。"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        sys.exit(f"[錯誤] xlsx 找不到 sheet「{SHEET_NAME}」")
    ws = wb[SHEET_NAME]

    by_need = {}
    xlsx_need_names = {}
    for row_idx in range(HEADER_ROW_COUNT + 1, ws.max_row + 1):
        need_id_raw = ws.cell(row=row_idx, column=COL_NEED_ID).value
        if need_id_raw is None or need_id_raw == "":
            continue  # 空列跳過

        try:
            need_id = int(need_id_raw)
        except (TypeError, ValueError):
            print(f"[警告] Row {row_idx} 的 need_id 不是整數：{need_id_raw!r}，已跳過")
            continue

        idx_raw = ws.cell(row=row_idx, column=COL_INDICATOR_IDX).value
        try:
            idx = int(idx_raw)
        except (TypeError, ValueError):
            print(f"[警告] Row {row_idx} 的 indicator_index 不是整數：{idx_raw!r}，已跳過")
            continue

        weight_raw = ws.cell(row=row_idx, column=COL_WEIGHT).value
        try:
            weight = int(weight_raw)
        except (TypeError, ValueError):
            print(f"[警告] Row {row_idx} 的 weight 不是整數：{weight_raw!r}，預設為 1")
            weight = 1

        indicator = {
            "idx": idx,
            "text": clean_text(ws.cell(row=row_idx, column=COL_INDICATOR).value),
            "label": clean_text(ws.cell(row=row_idx, column=COL_LABEL).value),
            "weight": weight,
            "hover": clean_text(ws.cell(row=row_idx, column=COL_HOVER).value),
        }

        by_need.setdefault(need_id, []).append(indicator)

        # 記錄 xlsx 這邊的 need 名稱（用於之後對比警告）
        name = clean_text(ws.cell(row=row_idx, column=COL_NEED).value)
        if name and need_id not in xlsx_need_names:
            xlsx_need_names[need_id] = name

    # 每個 need 內按 idx 排序，確保輸出穩定
    for nid in by_need:
        by_need[nid].sort(key=lambda x: x["idx"])

    return by_need, xlsx_need_names


def load_existing_json(json_path):
    """讀現有的 twa-matrix.json。"""
    if not json_path.exists():
        sys.exit(f"[錯誤] 找不到 {json_path}")
    with open(json_path, "r", encoding="utf-8") as f:
        return json.load(f)


def diff_indicator(old, new):
    """比對兩個 indicator，回傳有變動的欄位列表。"""
    changes = []
    for key in ("text", "label", "weight", "hover"):
        if old.get(key) != new.get(key):
            changes.append(key)
    return changes


# ---------------------------------------------------------------------------
# 主流程
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--xlsx", default=str(DEFAULT_XLSX), help=f"xlsx 路徑（預設 {DEFAULT_XLSX}）")
    parser.add_argument("--json", default=str(DEFAULT_JSON), help=f"JSON 路徑（預設 {DEFAULT_JSON}）")
    parser.add_argument("--dry-run", action="store_true", help="只印變動，不實際寫檔")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    json_path = Path(args.json)

    if not xlsx_path.exists():
        sys.exit(f"[錯誤] 找不到 {xlsx_path}")

    print(f"[讀取] {xlsx_path}")
    by_need, xlsx_need_names = read_xlsx_indicators(xlsx_path)

    print(f"[讀取] {json_path}")
    existing = load_existing_json(json_path)

    # 以現有 JSON 為骨架，只換 indicators
    output = []
    total_changes = 0
    per_need_changes = []

    for item in existing:
        nid = item["id"]
        new_item = {
            "id": nid,
            "v": item.get("v", ""),
            "n": item.get("n", ""),
            "d": item.get("d", ""),
        }

        # need 名稱不一致警告
        xlsx_name = xlsx_need_names.get(nid)
        if xlsx_name and xlsx_name != item.get("n"):
            print(f"[警告] need_id={nid} 名稱不一致 — JSON: {item.get('n')!r} / xlsx: {xlsx_name!r}（保留 JSON）")

        new_indicators = by_need.get(nid, [])
        if not new_indicators:
            print(f"[警告] xlsx 裡找不到 need_id={nid} 的任何指標，保留舊資料")
            new_item["indicators"] = item.get("indicators", [])
            output.append(new_item)
            continue

        # 對比每筆 indicator 的變動
        old_by_idx = {ind["idx"]: ind for ind in item.get("indicators", [])}
        need_changes = 0
        for new_ind in new_indicators:
            old_ind = old_by_idx.get(new_ind["idx"])
            if old_ind is None:
                need_changes += 1
                total_changes += 1
            else:
                changed_keys = diff_indicator(old_ind, new_ind)
                if changed_keys:
                    need_changes += 1
                    total_changes += 1

        if need_changes:
            per_need_changes.append((nid, item.get("n"), need_changes, len(new_indicators)))

        new_item["indicators"] = new_indicators
        output.append(new_item)

    # 變動摘要
    print()
    print("=" * 60)
    print("變動摘要")
    print("=" * 60)
    if per_need_changes:
        for nid, nname, changed, total in per_need_changes:
            print(f"  [{nid:>2}] {nname}：{changed}/{total} 筆 indicator 有變動")
    else:
        print("  （沒有任何 indicator 變動）")
    print()
    print(f"總計 {total_changes} 筆 indicator 變動 / 共 {sum(len(it['indicators']) for it in output)} 筆")

    # 寫檔
    if args.dry_run:
        print()
        print("[dry-run] 未寫檔。")
        return

    if total_changes == 0:
        print()
        print("沒有變動，跳過寫檔。")
        return

    with open(json_path, "w", encoding="utf-8", newline="\n") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
        f.write("\n")  # 尾端換行，Git 友善

    print()
    print(f"[完成] 已寫回 {json_path}")
    print("下一步：firebase deploy --only hosting")


if __name__ == "__main__":
    main()
